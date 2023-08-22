import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
import threading
import time


def update_table():
    global tree
    while True:
        # тут можно изменить название файла, при замене название также необходимо сменить в модуле main.py
        wb = load_workbook('биржевые данные.xlsx')
        sheet = wb.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(row)
        tree.delete(*tree.get_children())
        for row in rows:
            tree.insert("", "end", values=row)
        time.sleep(10)


def create_gui():
    global tree
    root = tk.Tk()
    root.title("Excel Table Viewer")

    tree = ttk.Treeview(root)
    tree["columns"] = ("#1", "#2", "#3", "#4", "#5", "#6", "#7", "#8")

    for i in range(8):
        tree.column(f"#{i + 1}", width=100, anchor='center')
        tree.heading(f"#{i + 1}", text=f"Column {i + 1}")

    tree.pack()

    update_thread = threading.Thread(target=update_table)
    update_thread.daemon = True
    update_thread.start()

    root.mainloop()


